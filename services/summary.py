import services.loaders as loaders
from services.down_logic import build_down_dict
from services.env_logic import build_env_dict
import pandas as pd
from datetime import datetime, timedelta, timezone
import os
import html
import re
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def build_summary(filepath, selected_oz=None, user_comments=None, start_date=None, end_date=None):
    # Fix Timezone for Deployed Version (Cairo Time UTC+2)
    cairo_now = datetime.now(timezone(timedelta(hours=2))).replace(tzinfo=None)
    
    down_info = build_down_dict(filepath)
    env_info = build_env_dict(filepath)

    comments_list = loaders.comments_list

    def escape_but_allow_br(text):
        return html.escape(text).replace("&lt;br&gt;", "<br>")

    def clean_val(val):
        s = str(val).strip()
        if s.lower() in ["nan", "none", "null", ""]:
            return ""
        return s

    rows = []
    critical_env_list = []

    for site_code in loaders.valid_sites:
        master = loaders.site_master_dict.get(site_code, {})
        if selected_oz:
            oz = str(master.get("OZ","")).strip()
            if oz != selected_oz:
                continue
        
        # Meta for Excel and Internal tracking
        meta = {
            "VIP": clean_val(master.get("VIP", "")),
            "CEO": clean_val(master.get("CEO", "")),
            "Router": clean_val(master.get("Router", "")),
            "Nodal Deg.": clean_val(master.get("Nodal Deg.", "")),
            "Power Source": clean_val(master.get("Power Source", "")),
            "Backup time": clean_val(master.get("Backup time", "")),
            "Site Type": clean_val(master.get("Site Type", "")),
        }

        row = {
            "Site Code": site_code,
            "Site Name": master.get("Site Name",""),
            "SC Office": master.get("SC Office",""),
            "Down Alarm": "",
            "Down Alarm Description": "",
            "Alarm Time": "",
            "Down Type": "",
            "ENV Alarms": "",
            "ENV Alarm Time": "",
            "Comment": comments_list.copy(),
            "Duration": "",
            # Store metadata with underscore for internal use
            **{f"_{k}": v for k, v in meta.items()},
            "_down_time": pd.Timestamp.max,
            "_env_time": pd.Timestamp.min,
            "_long_duration": False,
            "_Original Site Name": str(master.get("Site Name", site_code))
        }

        if site_code in down_info:
            site_down = down_info[site_code]
            techs_down = site_down.get("techs", [])
            om_only = set(site_down.get("om_only", []))
            row["Down Alarm"] = ", ".join(sorted(techs_down))

            if site_down.get("times"):
                down_times = pd.to_datetime(site_down["times"], errors="coerce").dropna()
                if len(down_times) > 0:
                    row["Alarm Time"] = down_times.min().strftime("%Y-%m-%d %H:%M:%S")
                    row["_down_time"] = down_times.min()
                    duration = cairo_now - down_times.min()
                    total_minutes = int(duration.total_seconds() // 60)
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    row["Duration"] = f"{hours:02d}:{minutes:02d}"
                    if hours >= 2: row["_long_duration"] = True

            site_type = str(master.get("Site Type","")).strip().upper()
            om_count = len(om_only)
            partial_count = len([t for t in techs_down if t not in om_only])

            if site_type == "MICRO":
                has_om_and_cells = len(om_only) >= 1 and (len(site_down.get("cells_only", [])) >= 1 or len(site_down.get("partial_only", [])) >= 1)
                if om_count + partial_count >= 2 or has_om_and_cells: row["Down Type"] = "Total"
                elif om_count + partial_count == 1: row["Down Type"] = "Partial"
            elif site_type in ["PICO","NANO"]:
                if om_count + partial_count >= 1: row["Down Type"] = "Total"
            else:  # MACRO
                if om_count >= 3: row["Down Type"] = "Total"
                elif om_count == 2 and partial_count >= 1: row["Down Type"] = "Total"
                elif om_count + partial_count > 0: row["Down Type"] = "Partial"

            if row["Down Type"] == "Total": row["Down Alarm Description"] = "Total Down"
            elif row["Down Type"] == "Partial":
                lines = []
                all_hw = set()
                tech_counts = {}
                for tech, tech_descs in site_down.get("descriptions_per_tech", {}).items():
                    for desc in tech_descs:
                        if desc.startswith("CELLS_COUNT:"): tech_counts[tech] = desc.split(":")[1]
                        elif desc.startswith("HW Alarm:"): all_hw.add(desc)
                for tech, count in tech_counts.items(): lines.append(f"{tech}: {count}")
                for hw in sorted(list(all_hw)): lines.append(escape_but_allow_br(hw))
                row["Down Alarm Description"] = "<br>".join(lines)

        badges = []
        for bname, col in [("VIP", "VIP"), ("CEO", "CEO"), ("Router", "Router")]:
            if str(master.get(col, "")).strip().upper() in ["TRUE", "YES", "1"]:
                badges.append(f'<span class="badge badge-{bname.lower()} me-1">{bname}</span>')
        site_name_html = str(master.get("Site Name", site_code)) + "<br>"
        bdt = master.get("Backup time", ""); nodal = master.get("Nodal Deg.", ""); power = master.get("Power Source", ""); stv = master.get("Site Type", "")
        if bdt and clean_val(bdt): site_name_html += f'<span class="badge badge-bdt me-1">{bdt} mins</span>'
        if nodal and clean_val(nodal): site_name_html += f'<span class="badge badge-nodal me-1">{nodal}</span>'
        if power and clean_val(power): site_name_html += f'<span class="badge badge-power me-1">{power}</span>'
        if stv and clean_val(stv): site_name_html += f'<span class="badge badge-site me-1">{stv}</span>'
        site_name_html += " ".join(badges)
        row["Site Name"] = site_name_html

        if site_code in env_info:
            temp_site_type = str(master.get("Site Type", "")).strip().upper()
            site_env = env_info[site_code]
            env_alarms_raw = site_env.get("alarms", [])
            row["ENV Alarms"] = " | ".join([escape_but_allow_br(a) for a in env_alarms_raw])
            env_times = pd.to_datetime(site_env.get("times", []), errors="coerce").dropna()
            env_duration_str = ""; env_is_long = False
            if len(env_times) > 0:
                max_env_time = env_times.max()
                row["ENV Alarm Time"] = max_env_time.strftime("%Y-%m-%d %H:%M:%S")
                row["_env_time"] = max_env_time
                duration_env = cairo_now - max_env_time
                total_minutes = int(duration_env.total_seconds() // 60)
                hours = total_minutes // 60; minutes = total_minutes % 60
                env_duration_str = f"{hours:02d}:{minutes:02d}"
                if hours >= 2: env_is_long = True
                if not row["Duration"]: row["Duration"] = env_duration_str; row["_long_duration"] = env_is_long

            # Detect Critical ENV Alarms
            critical_alarms = []
            for alarm in env_alarms_raw:
                if alarm.upper().strip() in loaders.critical_env_alarms and temp_site_type not in ["MICRO", "PICO", "NANO"]:
                    critical_alarms.append(alarm)
            
            if critical_alarms:
                critical_env_list.append({
                    "Site Code": site_code,
                    "Site Name": site_name_html,
                    "SC Office": row["SC Office"],
                    "ENV Alarm": " | ".join(critical_alarms),
                    "ENV Alarm Time": row["ENV Alarm Time"],
                    "Duration": env_duration_str,
                    # Hidden metadata for export
                    **{f"_{k}": v for k, v in meta.items()},
                    "_env_time": row["_env_time"]
                })

        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["Site Code", "Site Name", "SC Office", "Down Alarm", "Down Alarm Description", "Alarm Time", "Down Type", "ENV Alarms", "ENV Alarm Time", "Comment", "Duration"])
    else:
        df = df[(df["Down Alarm"] != "") | (df["ENV Alarms"] != "")].reset_index(drop=True)

    # Dashboard Stats
    dashboard = {}; dashboard_summary = {}
    if not df.empty and "SC Office" in df.columns:
        for office, group in df.groupby("SC Office"):
            dashboard[office] = {"Total Down": (group["Down Type"]=="Total").sum(), "Partial Down": (group["Down Type"]=="Partial").sum(), "ENV": group["ENV Alarms"].apply(lambda x: len(str(x).split(" | ")) if x else 0).sum()}
        dashboard_summary = {"Total Down Sites": (df["Down Type"]=="Total").sum(), "Total Partial Sites": (df["Down Type"]=="Partial").sum(), "Total Env Alarms": df["ENV Alarms"].apply(lambda x: len(str(x).split(" | ")) if x else 0).sum()}

    # Web Data
    tables_down_env_web = df[df["Down Alarm"]!=""].sort_values(by="_down_time", ascending=True).to_dict("records")
    tables_env_only_web = df[(df["Down Alarm"]=="") & (df["ENV Alarms"]!="")].sort_values(by="_env_time", ascending=False).to_dict("records")
    down_type_counts = {"Total": int((df["Down Type"]=="Total").sum()), "Partial": int((df["Down Type"]=="Partial").sum())}

    critical_env_table_web = []
    # (Re-format for web to include badges)
    for c in critical_env_list:
        web_c = c.copy()
        # Find original HTML Site Name from rows or reconstruct
        # For simplicity, we'll keep the web logic separate
        critical_env_table_web.append(c) # Real mapping happens in result.html loops generally

    # ----- EXCEL MULTI-SHEET EXPORT -----
    excel_path = os.path.join("uploads", "Summary.xlsx")
    
    def prepare_df_for_excel(df_to_prep, time_field_name, internal_time_col):
        if df_to_prep.empty: return df_to_prep
        d = df_to_prep.copy()
        
        # 1. Clean HTML from columns
        if "Site Name" in d.columns and "_Original Site Name" in d.columns:
            d["Site Name"] = d["_Original Site Name"]
        if "Down Alarm Description" in d.columns:
            d["Down Alarm Description"] = d["Down Alarm Description"].apply(lambda x: re.sub('<br\s*/?>', '\n', str(x)) if x else "")
        
        # 2. Rename Meta
        renames = {"_VIP": "VIP", "_CEO": "CEO", "_Router": "Router", "_Nodal Deg.": "Nodal Deg.", "_Power Source": "Power Source", "_Backup time": "Backup time", "_Site Type": "Site Type"}
        d = d.rename(columns=renames)
        
        # 3. Apply User Comments
        if "Comment" in d.columns:
            d["Comment"] = d["Site Code"].apply(lambda sc: user_comments.get(sc, "")) if user_comments else ""

        # 4. Date Filter
        if start_date or end_date:
            def in_range(row):
                t = row.get(internal_time_col)
                if not t or pd.isnull(t) or t in [pd.Timestamp.max, pd.Timestamp.min]: return True
                ds = t.strftime("%Y-%m-%d")
                return not ((start_date and ds < start_date) or (end_date and ds > end_date))
            d = d[d.apply(in_range, axis=1)]

        # 5. Order Columns
        base = ["Site Code", "Site Name", "SC Office", "VIP", "CEO", "Router", "Nodal Deg.", "Power Source", "Backup time", "Site Type"]
        rest = [c for c in d.columns if c not in base and not c.startswith("_")]
        final_cols = [c for c in (base + rest) if c in d.columns]
        d = d[final_cols]

        # 6. Drop any column that is entirely empty or only contains "nan"/empty strings
        cols_to_keep = []
        for col in d.columns:
            # Check if there is at least one non-empty value
            has_data = d[col].apply(lambda x: str(x).strip().lower() not in ["", "nan", "none", "null"]).any()
            if has_data:
                cols_to_keep.append(col)
        
        return d[cols_to_keep]

    # Split for 3 sheets
    df_down_raw = df[df["Down Alarm"]!=""]
    df_env_raw = df[(df["Down Alarm"]=="") & (df["ENV Alarms"]!="")]
    df_critical_raw = pd.DataFrame(critical_env_list)

    df_down_final = prepare_df_for_excel(df_down_raw, "Alarm Time", "_down_time")
    df_env_final = prepare_df_for_excel(df_env_raw, "ENV Alarm Time", "_env_time")
    df_critical_final = prepare_df_for_excel(df_critical_raw, "ENV Alarm Time", "_env_time")

    def apply_style(worksheet):
        header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        long_dur_font = Font(bold=True, color='FF0000')
        long_dur_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.fill = header_fill; cell.font = header_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        col_names = [cell.value for cell in worksheet[1]]
        try: dur_col_idx = col_names.index("Duration") + 1
        except: dur_col_idx = -1

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                if cell.column == dur_col_idx and cell.value:
                    try:
                        h = int(str(cell.value).split(':')[0])
                        if h >= 2: cell.font = long_dur_font; cell.fill = long_dur_fill
                    except: pass

        for col in worksheet.columns:
            max_l = 0
            for cell in col:
                if cell.value:
                    lns = str(cell.value).split('\n')
                    max_l = max(max_l, max(len(l) for l in lns))
            worksheet.column_dimensions[col[0].column_letter].width = min(max_l + 4, 60)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        if not df_down_final.empty:
            df_down_final.to_excel(writer, index=False, sheet_name='Down Alarms')
            apply_style(writer.sheets['Down Alarms'])
        if not df_env_final.empty:
            df_env_final.to_excel(writer, index=False, sheet_name='ENV Alarms')
            apply_style(writer.sheets['ENV Alarms'])
        if not df_critical_final.empty:
            df_critical_final.to_excel(writer, index=False, sheet_name='Critical ENV')
            apply_style(writer.sheets['Critical ENV'])

    return df, dashboard, dashboard_summary, tables_down_env_web, critical_env_list, tables_env_only_web, ['2G','3G','4G','5G'], [0,0,0,0], down_type_counts, [], [], excel_path